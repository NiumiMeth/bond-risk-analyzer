from __future__ import annotations

from datetime import date

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import auth
import io
import difflib

st.set_page_config(page_title="Portfolio Manager", layout="wide", page_icon="📈", initial_sidebar_state="expanded")

# require user or admin role to view portfolio page
#auth.require_role(["user", "admin"])

# ── DARK BANKING THEME ──────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

/* ── Base ── */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

.stApp {
    background: #0B0F1A;
    color: #E2E8F0;
}

/* ── Hide Streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2rem 2.5rem 4rem; max-width: 1600px; }

/* ── Top Header Bar ── */
.dash-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.5rem 0 1rem;
    border-bottom: 1px solid #1E2A3A;
    margin-bottom: 2rem;
}
.dash-logo {
    font-size: 1.1rem;
    font-weight: 600;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #94A3B8;
}
.dash-title {
    font-size: 1.6rem;
    font-weight: 600;
    color: #F1F5F9;
    letter-spacing: -0.02em;
}
.dash-badge {
    background: #0EA5E9;
    color: #fff;
    font-size: 0.7rem;
    font-weight: 600;
    padding: 3px 10px;
    border-radius: 20px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}

/* ── KPI Cards ── */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 1rem;
    margin-bottom: 1.8rem;
}
.kpi-card {
    background: #111827;
    border: 1px solid #1E2A3A;
    border-radius: 10px;
    padding: 1.25rem 1.5rem;
    position: relative;
    overflow: hidden;
}
.kpi-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #0EA5E9, #6366F1);
}
.kpi-label {
    font-size: 0.7rem;
    font-weight: 500;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #64748B;
    margin-bottom: 0.5rem;
}
.kpi-value {
    font-family: 'DM Mono', monospace;
    font-size: 1.25rem;
    font-weight: 500;
    color: #F1F5F9;
    letter-spacing: -0.02em;
}
.kpi-delta-pos {
    font-size: 0.78rem;
    color: #10B981;
    margin-top: 0.3rem;
    display: flex;
    align-items: center;
    gap: 3px;
}
.kpi-delta-neg {
    font-size: 0.78rem;
    color: #F43F5E;
    margin-top: 0.3rem;
    display: flex;
    align-items: center;
    gap: 3px;
}
.kpi-delta-neu {
    font-size: 0.78rem;
    color: #64748B;
    margin-top: 0.3rem;
}

/* ── Section Labels ── */
.section-label {
    font-size: 0.68rem;
    font-weight: 600;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #475569;
    margin-bottom: 0.75rem;
    margin-top: 0.25rem;
}

/* ── Panel Containers ── */
.panel {
    background: #111827;
    border: 1px solid #1E2A3A;
    border-radius: 10px;
    padding: 1.5rem;
    margin-bottom: 1rem;
}

/* ── Alert Banner ── */
.alert-banner {
    background: #1C1A10;
    border: 1px solid #854D0E;
    border-left: 3px solid #F59E0B;
    border-radius: 6px;
    padding: 0.6rem 1rem;
    font-size: 0.82rem;
    color: #FCD34D;
    margin-bottom: 1.2rem;
}

/* ── Tabs override ── */
.stTabs [data-baseweb="tab-list"] {
    background: transparent;
    gap: 0;
    border-bottom: 1px solid #1E2A3A;
    padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
    background: transparent;
    color: #64748B;
    font-size: 0.82rem;
    font-weight: 500;
    letter-spacing: 0.04em;
    padding: 0.6rem 1.2rem;
    border-bottom: 2px solid transparent;
    border-radius: 0;
}
.stTabs [aria-selected="true"] {
    background: transparent !important;
    color: #F1F5F9 !important;
    border-bottom: 2px solid #0EA5E9 !important;
}
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.5rem; }

/* ── Inputs ── */
.stDateInput > div > div, .stNumberInput > div > div, .stSelectbox > div > div {
    background: #0F172A !important;
    border-color: #1E2A3A !important;
    color: #E2E8F0 !important;
    border-radius: 6px !important;
}
.stTextInput input, .stNumberInput input {
    background: #0F172A !important;
    color: #E2E8F0 !important;
    border-color: #1E2A3A !important;
}

/* ── Buttons ── */
.stButton > button {
    background: #0EA5E9;
    color: #fff;
    border: none;
    border-radius: 6px;
    font-size: 0.82rem;
    font-weight: 500;
    padding: 0.5rem 1.2rem;
    letter-spacing: 0.04em;
    transition: background 0.15s;
}
.stButton > button:hover { background: #0284C7; }

/* ── Download button ── */
.stDownloadButton > button {
    background: transparent;
    border: 1px solid #1E2A3A;
    color: #94A3B8;
    font-size: 0.78rem;
    border-radius: 6px;
    padding: 0.4rem 1rem;
}
.stDownloadButton > button:hover { border-color: #0EA5E9; color: #0EA5E9; }

/* ── Dataframe ── */
.stDataFrame { border-radius: 8px; overflow: hidden; }
[data-testid="stDataFrameContainer"] {
    background: #0F172A;
    border: 1px solid #1E2A3A;
    border-radius: 8px;
}

/* ── Expander ── */
.streamlit-expanderHeader {
    background: #111827 !important;
    border: 1px solid #1E2A3A !important;
    border-radius: 8px !important;
    color: #94A3B8 !important;
    font-size: 0.84rem !important;
}
.streamlit-expanderContent {
    background: #111827 !important;
    border: 1px solid #1E2A3A !important;
    border-top: none !important;
}

/* ── Metric ── */
[data-testid="metric-container"] { background: transparent !important; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #080C14 !important;
    border-right: 1px solid #1E2A3A !important;
    min-width: 260px !important;
    width: 260px !important;
}
/* Always show sidebar expanded */
section[data-testid="stSidebar"] {
    transform: none !important;
    visibility: visible !important;
}
/* Style the collapse toggle so it blends with theme */
[data-testid="collapsedControl"] {
    background: #0B0F1A !important;
    border-right: 1px solid #1E2A3A !important;
    color: #475569 !important;
}
[data-testid="collapsedControl"]:hover {
    background: #111827 !important;
    color: #94A3B8 !important;
}

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #0B0F1A; }
::-webkit-scrollbar-thumb { background: #1E2A3A; border-radius: 4px; }
</style>
""", unsafe_allow_html=True)

EXPECTED_COLUMNS = [
    "Port. Index", "Instrument", "Deal No.", "ISIN",
    "Initial Inv Date", "Maturity Date", "Coupon",
    "Maturity Value", "YTM", "Yield", "Market value", "Duration",
]

# ── HELPER FUNCTIONS (unchanged) ────────────────────────────────────────────
def parse_number(value):
    if pd.isna(value): return np.nan
    text = str(value).strip().replace(",", "")
    if text == "": return np.nan
    try: return float(text)
    except ValueError: return np.nan

def parse_rate(value):
    if pd.isna(value): return np.nan
    text = str(value).strip()
    if text == "": return np.nan
    has_percent = "%" in text
    text = text.replace("%", "").replace(",", "")
    try: number = float(text)
    except ValueError: return np.nan
    if has_percent or number > 1: return number / 100.0
    return number

def clean_columns(df):
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]
    df = df.rename(columns={"Maturity Value ": "Maturity Value"})
    return df

def load_portfolio(uploaded_file):
    file_name = uploaded_file.name.lower()
    df = pd.read_csv(uploaded_file) if file_name.endswith(".csv") else pd.read_excel(uploaded_file)
    df = clean_columns(df)
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError("Missing columns: " + ", ".join(missing))
    df = df[EXPECTED_COLUMNS].copy()
    df = df.dropna(how="all")
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df = df[df["ISIN"].notna() & (df["ISIN"] != "") & (df["ISIN"] != "nan")].copy()
    df["Initial Inv Date"] = pd.to_datetime(df["Initial Inv Date"], dayfirst=True, errors="coerce")
    df["Maturity Date"] = pd.to_datetime(df["Maturity Date"], dayfirst=True, errors="coerce")
    for col in ["Maturity Value", "Market value", "Duration"]:
        df[col] = df[col].map(parse_number)
    df["Coupon"] = df["Coupon"].map(parse_rate)
    df["YTM"] = df["YTM"].map(parse_rate)
    df["Yield"] = df["Yield"].map(parse_rate)
    df = df.dropna(subset=["Initial Inv Date", "Maturity Date", "Maturity Value", "Coupon", "YTM", "Yield"])
    return df

def get_coupon_window(settlement_date, maturity_date, frequency=2):
    months = int(12 / frequency)
    settlement = pd.Timestamp(settlement_date).normalize()
    maturity = pd.Timestamp(maturity_date).normalize()
    if settlement >= maturity:
        raise ValueError("Settlement date must be earlier than maturity date.")
    next_coupon = maturity
    while True:
        prev_coupon = next_coupon - pd.DateOffset(months=months)
        if prev_coupon <= settlement < next_coupon:
            return prev_coupon, next_coupon
        next_coupon = prev_coupon

def excel_price_actual_actual(settlement_date, maturity_date, coupon_rate, annual_yield, redemption=100.0, frequency=2):
    settlement = pd.Timestamp(settlement_date).normalize()
    maturity = pd.Timestamp(maturity_date).normalize()
    if settlement >= maturity: return 0.0, 0.0, 0.0
    prev_coupon, next_coupon = get_coupon_window(settlement, maturity, frequency)
    e = (next_coupon - prev_coupon).days
    a = (settlement - prev_coupon).days
    dsc = (next_coupon - settlement).days
    if e <= 0: return 0.0, 0.0, 0.0
    coupon_per_100 = 100.0 * coupon_rate / frequency
    discount_base = max(1.0 + annual_yield / frequency, 1e-8)
    n = 0
    current = maturity
    while current > settlement:
        n += 1
        current = current - pd.DateOffset(months=int(12 / frequency))
    if n <= 0: return 0.0, 0.0, 0.0
    exponent = (n - 1) + (dsc / e)
    pv_red_plus_coupon = (redemption + coupon_per_100) / (discount_base ** exponent)
    pv_intermediate = sum(coupon_per_100 / (discount_base ** ((k - 1) + (dsc / e))) for k in range(1, n))
    accrued_100 = coupon_per_100 * (a / e)
    clean_price_100 = pv_red_plus_coupon + pv_intermediate - accrued_100
    full_price_100 = clean_price_100 + accrued_100
    return clean_price_100, accrued_100, full_price_100

def run_portfolio_valuation(df, valuation_date):
    output = df.copy()
    clean_price_100, accrued_100, price_100 = [], [], []
    clean_value, full_value, initial_inv_value = [], [], []
    book_value, gain_loss = [], []
    for _, row in output.iterrows():
        face = float(row["Maturity Value"])
        purchase_date = pd.Timestamp(row["Initial Inv Date"])
        maturity_date = pd.Timestamp(row["Maturity Date"])
        coupon_rate = float(row["Coupon"])
        purchased_ytm = float(row["YTM"])
        selling_ytm = float(row["Yield"])
        clean_100_now, accrued_now_100, _ = excel_price_actual_actual(valuation_date, maturity_date, coupon_rate, selling_ytm, 100.0, 2)
        clean_100_now = round(clean_100_now, 4)
        accrued_now_100 = round(accrued_now_100, 4)
        full_100_now = round(clean_100_now + accrued_now_100, 4)
        init_price_100, _, _ = excel_price_actual_actual(purchase_date, maturity_date, coupon_rate, purchased_ytm, 100.0, 2)
        init_price_100 = round(init_price_100, 4)
        init_value = init_price_100 * (face / 100.0)
        total_days = max((maturity_date - purchase_date).days, 1)
        elapsed_days = min(max((pd.Timestamp(valuation_date) - purchase_date).days, 0), total_days)
        book_val = (((face - init_value) / total_days) * elapsed_days) + init_value
        clean_val = clean_100_now * (face / 100.0)
        full_val = full_100_now * (face / 100.0)
        gl_value = clean_val - book_val
        clean_price_100.append(clean_100_now)
        accrued_100.append(accrued_now_100)
        price_100.append(full_100_now)
        clean_value.append(clean_val)
        full_value.append(full_val)
        initial_inv_value.append(init_value)
        book_value.append(book_val)
        gain_loss.append(gl_value)
    output["Clean Price"] = clean_price_100
    output["Accrued Int"] = accrued_100
    output["Price 100%"] = price_100
    output["Clean Value"] = clean_value
    output["Full Value"] = full_value
    output["Initial Inv Value"] = initial_inv_value
    output["Book Value"] = book_value
    output["Gain/Loss"] = gain_loss
    return output

# ── CHART HELPERS ────────────────────────────────────────────────────────────
CHART_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="DM Sans", color="#94A3B8", size=11),
    margin=dict(l=0, r=0, t=28, b=0),
    xaxis=dict(showgrid=False, zeroline=False, color="#475569"),
    yaxis=dict(showgrid=True, gridcolor="#1E2A3A", zeroline=False, color="#475569"),
    legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#94A3B8", size=10)),
)

def kpi_card(label, value, delta=None, delta_label="", positive_is_good=True):
    if delta is None:
        delta_html = f'<div class="kpi-delta-neu">—</div>'
    else:
        pos = delta >= 0
        good = pos if positive_is_good else not pos
        cls = "kpi-delta-pos" if good else "kpi-delta-neg"
        arrow = "▲" if pos else "▼"
        delta_html = f'<div class="{cls}">{arrow} {delta_label}</div>'
    return f"""
    <div class="kpi-card">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        {delta_html}
    </div>"""

def fmt_num(n, decimals=2):
    return f"{n:,.{decimals}f}"

def maturity_ladder_chart(df):
    df2 = df.copy()
    df2["Maturity Year"] = pd.to_datetime(df2["Maturity Date"]).dt.year
    ladder = df2.groupby("Maturity Year")["Maturity Value"].sum().reset_index()
    fig = go.Figure(go.Bar(
        x=ladder["Maturity Year"].astype(str),
        y=ladder["Maturity Value"],
        marker=dict(color="#0EA5E9", opacity=0.85),
        hovertemplate="<b>%{x}</b><br>%{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(**{k: v for k, v in CHART_LAYOUT.items()}, title=dict(text="Maturity Ladder", font=dict(size=12, color="#64748B")), height=240)
    return fig

def allocation_donut(df):
    data = df.groupby("ISIN")["Full Value"].sum().reset_index()
    fig = go.Figure(go.Pie(
        labels=data["ISIN"],
        values=data["Full Value"],
        hole=0.62,
        textinfo="none",
        hovertemplate="<b>%{label}</b><br>%{value:,.0f}<br>%{percent}<extra></extra>",
        marker=dict(colors=["#0EA5E9","#6366F1","#10B981","#F59E0B","#F43F5E","#8B5CF6","#14B8A6","#FB923C"],
                    line=dict(color="#0B0F1A", width=2)),
    ))
    fig.update_layout(
        **{k: v for k, v in CHART_LAYOUT.items() if k not in ("xaxis", "yaxis", "legend", "margin")},
        title=dict(text="Allocation by ISIN", font=dict(size=12, color="#64748B")),
        height=240,
        legend=dict(orientation="v", x=1.02, y=0.5, bgcolor="rgba(0,0,0,0)", font=dict(color="#94A3B8", size=9)),
        margin=dict(l=0, r=80, t=28, b=0),
    )
    return fig

def sensitivity_curve(df_clean, valuation_ts):
    shocks = [s / 100 for s in range(-200, 210, 25)]
    values = []
    for shock in shocks:
        tmp = df_clean.copy()
        tmp["Yield"] = tmp["Yield"].fillna(0) + shock
        val = run_portfolio_valuation(tmp, valuation_ts)
        values.append(float(val["Full Value"].sum()))
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=[s * 100 for s in shocks], y=values,
        mode="lines+markers",
        line=dict(color="#0EA5E9", width=2),
        marker=dict(size=5, color="#0EA5E9"),
        fill="tozeroy", fillcolor="rgba(14,165,233,0.07)",
        hovertemplate="<b>%{x:+.0f} bps</b><br>%{y:,.0f}<extra></extra>",
    ))
    fig.update_layout(**{k: v for k, v in CHART_LAYOUT.items() if k not in ("xaxis",)},
        title=dict(text="Portfolio Full Value vs Yield Shock", font=dict(size=12, color="#64748B")),
        xaxis=dict(title="Shock (bps)", showgrid=False, zeroline=True, zerolinecolor="#1E2A3A", color="#475569"),
        height=280)
    return fig

def yield_bar(df):
    fig = go.Figure()
    fig.add_trace(go.Bar(name="YTM", x=df["ISIN"], y=df["YTM"] * 100,
        marker_color="#6366F1", hovertemplate="%{x}<br>YTM: %{y:.2f}%<extra></extra>"))
    fig.add_trace(go.Bar(name="Yield", x=df["ISIN"], y=df["Yield"] * 100,
        marker_color="#0EA5E9", hovertemplate="%{x}<br>Yield: %{y:.2f}%<extra></extra>"))
    fig.update_layout(**{k: v for k, v in CHART_LAYOUT.items()},
        title=dict(text="YTM vs Current Yield", font=dict(size=12, color="#64748B")),
        barmode="group", height=240)
    return fig

# ── TEMPLATE GENERATION ──────────────────────────────────────────────────────
def make_template_excel():
    """Generate a perfectly formatted template Excel file with sample data + notes."""
    sample = pd.DataFrame([
        {
            "Port. Index": "PF-001",
            "Instrument": "T-Bond",
            "Deal No.": "DL-10001",
            "ISIN": "US912828ZT04",
            "Initial Inv Date": "15/01/2023",
            "Maturity Date": "15/01/2028",
            "Coupon": "5.50%",
            "Maturity Value": "1,000,000",
            "YTM": "5.25%",
            "Yield": "5.30%",
            "Market value": "998,500",
            "Duration": "4.32",
        },
        {
            "Port. Index": "PF-001",
            "Instrument": "Corp Bond",
            "Deal No.": "DL-10002",
            "ISIN": "XS1234567890",
            "Initial Inv Date": "20/03/2022",
            "Maturity Date": "20/03/2027",
            "Coupon": "4.75%",
            "Maturity Value": "500,000",
            "YTM": "4.50%",
            "Yield": "4.60%",
            "Market value": "501,200",
            "Duration": "3.87",
        },
    ])
    notes = pd.DataFrame([
        {"Column": "Port. Index",      "Format": "Text",         "Required": "Yes", "Notes": "Portfolio identifier, e.g. PF-001"},
        {"Column": "Instrument",       "Format": "Text",         "Required": "Yes", "Notes": "Bond type, e.g. T-Bond, Corp Bond, Sukuk"},
        {"Column": "Deal No.",         "Format": "Text",         "Required": "Yes", "Notes": "Unique deal/trade reference number"},
        {"Column": "ISIN",             "Format": "Text (12 chars)", "Required": "Yes", "Notes": "International Securities Identification Number"},
        {"Column": "Initial Inv Date", "Format": "DD/MM/YYYY",   "Required": "Yes", "Notes": "Settlement / purchase date"},
        {"Column": "Maturity Date",    "Format": "DD/MM/YYYY",   "Required": "Yes", "Notes": "Bond maturity date"},
        {"Column": "Coupon",           "Format": "% e.g. 5.50%", "Required": "Yes", "Notes": "Annual coupon rate"},
        {"Column": "Maturity Value",   "Format": "Number",       "Required": "Yes", "Notes": "Face/par value at maturity"},
        {"Column": "YTM",              "Format": "% e.g. 5.25%", "Required": "Yes", "Notes": "Yield to maturity at purchase"},
        {"Column": "Yield",            "Format": "% e.g. 5.30%", "Required": "Yes", "Notes": "Current market yield"},
        {"Column": "Market value",     "Format": "Number",       "Required": "Yes", "Notes": "Current market value"},
        {"Column": "Duration",         "Format": "Number",       "Required": "Yes", "Notes": "Modified duration in years"},
    ])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sample.to_excel(writer, sheet_name="Portfolio", index=False)
        notes.to_excel(writer, sheet_name="Column Guide", index=False)
        # Style the Portfolio sheet
        wb = writer.book
        ws_port = writer.sheets["Portfolio"]
        ws_notes = writer.sheets["Column Guide"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_font = Font(color="94A3B8", bold=True, size=10)
        note_fill   = PatternFill("solid", fgColor="1E3A5F")
        for ws in [ws_port, ws_notes]:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        for col in ws_port.columns:
            ws_port.column_dimensions[col[0].column_letter].width = 18
        for col in ws_notes.columns:
            ws_notes.column_dimensions[col[0].column_letter].width = 22
        # Highlight required column
        req_fill = PatternFill("solid", fgColor="14532D")
        for row in ws_notes.iter_rows(min_row=2):
            if row[2].value == "Yes":
                for cell in row:
                    cell.fill = req_fill
    buf.seek(0)
    return buf.getvalue()


# ── COLUMN AUTO-MAPPER ────────────────────────────────────────────────────────
COLUMN_ALIASES = {
    "Port. Index":      ["port index", "portfolio index", "port idx", "portfolio", "port"],
    "Instrument":       ["instrument", "instr", "security type", "type", "bond type"],
    "Deal No.":         ["deal no", "deal no.", "deal number", "deal_no", "trade no", "trade number", "ref"],
    "ISIN":             ["isin", "isin code", "security id", "bond id"],
    "Initial Inv Date": ["initial inv date", "investment date", "purchase date", "settlement date",
                         "inv date", "trade date", "buy date", "initial date"],
    "Maturity Date":    ["maturity date", "mat date", "maturity", "expiry date", "redemption date"],
    "Coupon":           ["coupon", "coupon rate", "coupon %", "rate", "interest rate"],
    "Maturity Value":   ["maturity value", "face value", "par value", "nominal", "principal",
                         "face", "par", "notional", "maturity value "],
    "YTM":              ["ytm", "yield to maturity", "purchase ytm", "ytm %"],
    "Yield":            ["yield", "current yield", "market yield", "yield %", "selling yield"],
    "Market value":     ["market value", "market val", "mkt value", "mv", "price", "market price"],
    "Duration":         ["duration", "modified duration", "mod duration", "dur"],
}

def auto_map_columns(file_cols: list[str]) -> dict[str, str]:
    """Return {required_col: matched_file_col} using alias + fuzzy matching."""
    file_lower = {c.strip().lower(): c for c in file_cols}
    mapping = {}
    for req_col, aliases in COLUMN_ALIASES.items():
        # 1. Exact match (case-insensitive)
        for alias in aliases:
            if alias in file_lower:
                mapping[req_col] = file_lower[alias]
                break
        # 2. Fuzzy match if no exact hit
        if req_col not in mapping:
            all_aliases = aliases + [req_col.lower()]
            matches = difflib.get_close_matches(
                req_col.lower(), list(file_lower.keys()), n=1, cutoff=0.72
            )
            if matches:
                mapping[req_col] = file_lower[matches[0]]
    return mapping


def load_raw_file(uploaded_file) -> pd.DataFrame:
    """Read raw file without any column expectations."""
    name = uploaded_file.name.lower()
    uploaded_file.seek(0)
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    else:
        return pd.read_excel(uploaded_file)


def apply_mapping_and_load(raw_df: pd.DataFrame, col_map: dict[str, str]) -> pd.DataFrame:
    """Rename columns per mapping, then run full load_portfolio logic."""
    reverse = {v: k for k, v in col_map.items()}
    df = raw_df.rename(columns=reverse)
    df = clean_columns(df)
    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError("Still missing columns after mapping: " + ", ".join(missing))
    df = df[EXPECTED_COLUMNS].copy()
    df = df.dropna(how="all")
    df["ISIN"] = df["ISIN"].astype(str).str.strip()
    df = df[df["ISIN"].notna() & (df["ISIN"] != "") & (df["ISIN"] != "nan")].copy()
    df["Initial Inv Date"] = pd.to_datetime(df["Initial Inv Date"], dayfirst=True, errors="coerce")
    df["Maturity Date"]    = pd.to_datetime(df["Maturity Date"],    dayfirst=True, errors="coerce")
    for col in ["Maturity Value", "Market value", "Duration"]:
        df[col] = df[col].map(parse_number)
    df["Coupon"] = df["Coupon"].map(parse_rate)
    df["YTM"]    = df["YTM"].map(parse_rate)
    df["Yield"]  = df["Yield"].map(parse_rate)
    df = df.dropna(subset=["Initial Inv Date", "Maturity Date", "Maturity Value", "Coupon", "YTM", "Yield"])
    return df


# ── UPLOAD PANEL ──────────────────────────────────────────────────────────────
def render_upload_panel():
    """Full guided upload experience: template -> upload -> map -> preview -> confirm."""

    # Template download
    template_bytes = make_template_excel()
    st.download_button(
        label="📥 Download Template",
        data=template_bytes,
        file_name="portfolio_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Download a pre-formatted Excel template with sample data and column guide",
    )

    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    uploaded_file = st.file_uploader(
        "Drop your CSV or Excel file here",
        type=["csv", "xlsx", "xls"],
        label_visibility="visible",
        help="Accepted formats: .csv, .xlsx, .xls — columns will be auto-detected",
    )

    if uploaded_file is None:
        # Show format hint when no file loaded
        if st.session_state["portfolio_df"].empty:
            st.markdown("""
            <div style="background:#0A1628;border:1px solid #1E2A3A;border-radius:8px;padding:.75rem 1rem;margin-top:.5rem;">
                <div style="font-size:.68rem;font-weight:600;letter-spacing:.1em;text-transform:uppercase;color:#334155;margin-bottom:.5rem;">Required Columns</div>
                <div style="display:flex;flex-wrap:wrap;gap:4px;">
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Port. Index</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Instrument</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Deal No.</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">ISIN</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Initial Inv Date</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Maturity Date</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Coupon</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Maturity Value</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">YTM</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Yield</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Market value</span>
                    <span style="background:#0F172A;color:#64748B;border:1px solid #1E2A3A;border-radius:4px;padding:2px 7px;font-size:.7rem;font-family:monospace;">Duration</span>
                </div>
                <div style="font-size:.7rem;color:#334155;margin-top:.5rem;">Column names are auto-detected — exact match not required.</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            # Portfolio loaded — show status + clear button
            n = len(st.session_state["portfolio_df"])
            isins = st.session_state["portfolio_df"]["ISIN"].nunique() if "ISIN" in st.session_state["portfolio_df"].columns else 0
            st.markdown(f"""
            <div style="background:#0A1628;border:1px solid #1E2A3A;border-radius:8px;padding:.75rem 1rem;margin-top:.5rem;">
                <div style="display:flex;justify-content:space-between;align-items:center;">
                    <div>
                        <div style="font-size:.68rem;color:#475569;text-transform:uppercase;letter-spacing:.08em;">Loaded</div>
                        <div style="font-family:'DM Mono',monospace;color:#10B981;font-size:.95rem;font-weight:600;">{n} holdings · {isins} ISINs</div>
                    </div>
                    <div style="width:8px;height:8px;border-radius:50%;background:#10B981;box-shadow:0 0 6px #10B981;"></div>
                </div>
            </div>
            """, unsafe_allow_html=True)
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
            if st.button("🗑️ Clear Portfolio", use_container_width=True):
                st.session_state["portfolio_df"] = pd.DataFrame(columns=EXPECTED_COLUMNS)
                st.session_state.pop("_upload_mapping", None)
                st.session_state.pop("_upload_raw", None)
                st.session_state.pop("_upload_preview", None)
                st.rerun()
        return

    # ── File uploaded: run auto-detection ──
    try:
        raw_df = load_raw_file(uploaded_file)
    except Exception as e:
        st.error(f"Could not read file: {e}")
        return

    if raw_df.empty:
        st.error("File appears to be empty.")
        return

    file_cols = list(raw_df.columns)
    auto_map  = auto_map_columns(file_cols)

    # Check which required columns were NOT auto-mapped
    unmapped = [c for c in EXPECTED_COLUMNS if c not in auto_map]
    mapped   = [c for c in EXPECTED_COLUMNS if c in auto_map]

    # ── Show mapping summary ──
    all_perfect = len(unmapped) == 0
    if all_perfect:
        st.markdown(f"""
        <div style="background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.25);
                    border-radius:8px;padding:.6rem .9rem;margin:.4rem 0;font-size:.78rem;color:#10B981;">
            ✓ All {len(EXPECTED_COLUMNS)} columns detected automatically
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div style="background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.25);
                    border-radius:8px;padding:.6rem .9rem;margin:.4rem 0;font-size:.78rem;color:#F59E0B;">
            ⚠ {len(mapped)} of {len(EXPECTED_COLUMNS)} columns matched — {len(unmapped)} need manual mapping
        </div>""", unsafe_allow_html=True)

        # Manual mapping UI for unmapped columns
        st.markdown('<div style="font-size:.7rem;color:#475569;margin:.5rem 0 .25rem;">Map remaining columns:</div>', unsafe_allow_html=True)
        file_col_options = ["— skip —"] + file_cols
        for req_col in unmapped:
            chosen = st.selectbox(
                f"{req_col}",
                options=file_col_options,
                key=f"map_{req_col}",
                help=f"Which column in your file corresponds to '{req_col}'?",
            )
            if chosen != "— skip —":
                auto_map[req_col] = chosen

    # Still missing after manual mapping?
    still_missing = [c for c in EXPECTED_COLUMNS if c not in auto_map]
    if still_missing:
        st.markdown(f"""
        <div style="background:rgba(244,63,94,0.08);border:1px solid rgba(244,63,94,0.25);
                    border-radius:8px;padding:.6rem .9rem;margin:.4rem 0;font-size:.75rem;color:#F87171;">
            Missing: {", ".join(still_missing)}<br>
            <span style="color:#64748B;">Map or rename these columns, then re-upload.</span>
        </div>""", unsafe_allow_html=True)
        return

    # ── Build preview ──
    try:
        preview_df = apply_mapping_and_load(raw_df, auto_map)
    except Exception as e:
        st.error(f"Processing error: {e}")
        return

    total_raw   = len(raw_df)
    total_valid = len(preview_df)
    dropped     = total_raw - total_valid

    # ── Validation summary ──
    st.markdown(f"""
    <div style="background:#0A1628;border:1px solid #1E2A3A;border-radius:8px;
                padding:.75rem 1rem;margin:.5rem 0;">
        <div style="font-size:.68rem;font-weight:600;letter-spacing:.1em;
                    text-transform:uppercase;color:#334155;margin-bottom:.5rem;">Preview</div>
        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:.5rem;text-align:center;">
            <div>
                <div style="font-family:'DM Mono',monospace;font-size:1.1rem;
                            font-weight:600;color:#F1F5F9;">{total_raw}</div>
                <div style="font-size:.68rem;color:#475569;">Rows in file</div>
            </div>
            <div>
                <div style="font-family:'DM Mono',monospace;font-size:1.1rem;
                            font-weight:600;color:#10B981;">{total_valid}</div>
                <div style="font-size:.68rem;color:#475569;">Valid holdings</div>
            </div>
            <div>
                <div style="font-family:'DM Mono',monospace;font-size:1.1rem;
                            font-weight:600;color:{"#F59E0B" if dropped > 0 else "#475569"};">{dropped}</div>
                <div style="font-size:.68rem;color:#475569;">Rows skipped</div>
            </div>
        </div>
        {f'<div style="font-size:.7rem;color:#F59E0B;margin-top:.5rem;border-top:1px solid #1E2A3A;padding-top:.4rem;">'
          f'{dropped} row(s) skipped — missing required fields (dates, rates, or face value). '
          f'Download the template to see expected formats.</div>' if dropped > 0 else ""}
    </div>
    """, unsafe_allow_html=True)

    if preview_df.empty:
        st.error("No valid rows found after processing. Check your file format.")
        return

    # Sample preview table (first 3 rows, key columns only)
    preview_cols = ["ISIN", "Deal No.", "Maturity Date", "Coupon", "Maturity Value", "Yield"]
    available_preview = [c for c in preview_cols if c in preview_df.columns]
    st.dataframe(
        preview_df[available_preview].head(3).style.format({
            "Maturity Value": "{:,.0f}",
            "Coupon": "{:.2%}",
            "Yield": "{:.2%}",
        }, na_rep="—"),
        use_container_width=True,
        height=130,
    )

    # ── Confirm button ──
    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
    if st.button(f"✓ Load {total_valid} Holdings", use_container_width=True, type="primary"):
        existing = st.session_state["portfolio_df"]
        combined = pd.concat([existing, preview_df], ignore_index=True)
        combined = combined.drop_duplicates(subset=["ISIN", "Deal No."], keep="last").reset_index(drop=True)
        st.session_state["portfolio_df"] = combined
        st.success(f"✓ {total_valid} holdings loaded successfully")
        st.rerun()


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    # ── Header ──
    st.markdown("""
    <div class="dash-header">
        <div class="dash-logo">Fixed Income</div>
        <div class="dash-title">Bond Portfolio Manager</div>
        <div class="dash-badge">Live</div>
    </div>
    """, unsafe_allow_html=True)

    if "portfolio_df" not in st.session_state:
        st.session_state["portfolio_df"] = pd.DataFrame(columns=EXPECTED_COLUMNS)

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        # Extra sidebar styles
        st.markdown("""
        <style>
        [data-testid="stSidebar"] { padding-top: 0 !important; }
        [data-testid="stSidebar"] > div:first-child { padding: 1.2rem 1rem 2rem; }

        /* Nav links */
        .nav-section { margin-bottom: 1.4rem; }
        .nav-label {
            font-size: 0.62rem;
            font-weight: 600;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            color: #334155;
            margin-bottom: 0.4rem;
            padding-left: 4px;
        }
        .nav-link {
            display: flex;
            align-items: center;
            gap: 8px;
            padding: 0.5rem 0.75rem;
            border-radius: 7px;
            font-size: 0.83rem;
            font-weight: 500;
            color: #94A3B8;
            cursor: pointer;
            transition: background 0.12s, color 0.12s;
            text-decoration: none;
            margin-bottom: 2px;
        }
        .nav-link:hover { background: #1E2A3A; color: #E2E8F0; }
        .nav-link.active { background: rgba(14,165,233,0.12); color: #0EA5E9; }
        .nav-link .icon { font-size: 0.9rem; width: 18px; text-align: center; }

        /* Divider */
        .sb-divider {
            border: none;
            border-top: 1px solid #1E2A3A;
            margin: 1rem 0;
        }

        /* Portfolio status pill */
        .port-status {
            display: flex;
            align-items: center;
            justify-content: space-between;
            background: #0F172A;
            border: 1px solid #1E2A3A;
            border-radius: 8px;
            padding: 0.6rem 0.9rem;
            margin-bottom: 1rem;
        }
        .port-status-label { font-size: 0.72rem; color: #475569; }
        .port-status-value { font-size: 0.82rem; font-weight: 600; color: #E2E8F0; font-family: 'DM Mono', monospace; }

        /* Logout button */
        [data-testid="stSidebar"] .stButton > button {
            background: transparent !important;
            border: 1px solid #1E2A3A !important;
            color: #64748B !important;
            font-size: 0.78rem !important;
            width: 100% !important;
            border-radius: 7px !important;
            padding: 0.4rem !important;
        }
        [data-testid="stSidebar"] .stButton > button:hover {
            border-color: #F43F5E !important;
            color: #F43F5E !important;
            background: rgba(244,63,94,0.06) !important;
        }

        /* File uploader */
        [data-testid="stSidebar"] [data-testid="stFileUploader"] {
            background: #0F172A;
            border: 1px dashed #1E2A3A;
            border-radius: 8px;
            padding: 0.5rem;
        }

        /* Success/error in sidebar */
        [data-testid="stSidebar"] .stAlert { font-size: 0.78rem !important; }
        </style>
        """, unsafe_allow_html=True)

        # ── User profile card ──
        auth.render_sidebar_user_panel()

        # ── Navigation ──
        st.markdown("""
        <div class="nav-section">
            <div class="nav-label">Navigation</div>
            <div class="nav-link active"><span class="icon">📊</span> Portfolio Dashboard</div>
            <div class="nav-link"><span class="icon">📋</span> Holdings</div>
            <div class="nav-link"><span class="icon">📈</span> Scenario Analysis</div>
            <div class="nav-link"><span class="icon">🔍</span> Bond Deep-Dive</div>
        </div>
        <hr class="sb-divider"/>
        """, unsafe_allow_html=True)

        # ── Valuation date ──
        st.markdown('<div class="nav-label">Valuation Date</div>', unsafe_allow_html=True)
        valuation_date = st.date_input("", value=date.today(), label_visibility="collapsed")

        # ── Portfolio status summary ──
        if not st.session_state["portfolio_df"].empty:
            n = len(st.session_state["portfolio_df"])
            st.markdown(f"""
            <div class="port-status">
                <span class="port-status-label">Holdings loaded</span>
                <span class="port-status-value">{n}</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<hr class="sb-divider"/>', unsafe_allow_html=True)

        # ── Data upload (open in main) ──
        st.markdown('<div class="nav-label">Data Upload</div>', unsafe_allow_html=True)
        if st.button("📁 Upload Portfolio", key="open_upload_in_main"):
            st.session_state["_show_upload_in_main"] = True

    valuation_timestamp = pd.Timestamp(valuation_date)

    # ── Show upload panel in main when portfolio is empty or when requested from sidebar
    show_upload = st.session_state.get("_show_upload_in_main", False) or st.session_state["portfolio_df"].empty
    if show_upload:
        render_upload_panel()
        # If portfolio still empty, stop here so user can interact with the uploader
        if st.session_state["portfolio_df"].empty:
            return
        # Otherwise clear the flag and continue to render dashboard
        st.session_state.pop("_show_upload_in_main", None)

    # ── Compute ──
    try:
        cleaned = clean_columns(st.session_state["portfolio_df"]).copy()
        for col in ["Maturity Value", "Market value", "Duration"]:
            cleaned[col] = cleaned[col].map(parse_number)
        cleaned["Coupon"] = cleaned["Coupon"].map(parse_rate)
        cleaned["YTM"] = cleaned["YTM"].map(parse_rate)
        cleaned["Yield"] = cleaned["Yield"].map(parse_rate)

        valued = run_portfolio_valuation(cleaned, valuation_timestamp)

        total_book = float(valued["Book Value"].sum())
        total_full = float(valued["Full Value"].sum())
        total_clean = float(valued["Clean Value"].sum())
        total_gl = float(valued["Gain/Loss"].sum())
        gl_pct = (total_gl / total_book * 100) if total_book else 0.0
        n_holdings = len(valued)

        # ── Maturity alert ──
        today = pd.Timestamp(valuation_date)
        maturing_soon = valued[
            (pd.to_datetime(valued["Maturity Date"]) - today).dt.days.between(0, 90)
        ]
        if not maturing_soon.empty:
            st.markdown(
                f'<div class="alert-banner">⚠️ {len(maturing_soon)} bond(s) maturing within 90 days — '
                f'Face value: {fmt_num(float(maturing_soon["Maturity Value"].sum()))}</div>',
                unsafe_allow_html=True,
            )

        # ── KPI Cards ──
        gl_sign = "+" if total_gl >= 0 else ""
        kpi_html = f"""
        <div class="kpi-grid">
            {kpi_card("Total Book Value", fmt_num(total_book))}
            {kpi_card("Full Market Value", fmt_num(total_full),
                delta=total_full - total_book,
                delta_label=fmt_num(total_full - total_book),
                positive_is_good=True)}
            {kpi_card("Clean Value", fmt_num(total_clean))}
            {kpi_card("Unrealized Gain / Loss", f"{gl_sign}{fmt_num(total_gl)}",
                delta=total_gl,
                delta_label=f"{gl_sign}{gl_pct:.2f}% of book",
                positive_is_good=True)}
        </div>
        """
        st.markdown(kpi_html, unsafe_allow_html=True)

        # ── Secondary metrics row ──
        avg_coupon = float(cleaned["Coupon"].mean()) * 100
        avg_yield = float(cleaned["Yield"].mean()) * 100
        avg_ytm = float(cleaned["YTM"].mean()) * 100
        avg_dur = float(cleaned["Duration"].mean()) if "Duration" in cleaned.columns else 0.0

        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Holdings", f"{n_holdings}")
        s2.metric("Avg. Coupon", f"{avg_coupon:.2f}%")
        s3.metric("Avg. Yield", f"{avg_yield:.2f}%")
        s4.metric("Avg. Duration", f"{avg_dur:.2f}")

        st.markdown("---")

        # ── Charts row ──
        c1, c2, c3 = st.columns([1.1, 1, 1.2])
        with c1:
            st.plotly_chart(maturity_ladder_chart(valued), use_container_width=True, config={"displayModeBar": False})
        with c2:
            st.plotly_chart(yield_bar(cleaned), use_container_width=True, config={"displayModeBar": False})
        with c3:
            st.plotly_chart(allocation_donut(valued), use_container_width=True, config={"displayModeBar": False})

        st.markdown("---")

        # ── Tabs ──
        tab_holdings, tab_scenario, tab_deepdive = st.tabs([
            "  Current Holdings  ",
            "  Scenario Analysis  ",
            "  Bond Deep-Dive  ",
        ])

        fmt = {
            "Maturity Value": "{:,.0f}",
            "Market value": "{:,.2f}",
            "Initial Inv Value": "{:,.2f}",
            "Clean Value": "{:,.2f}",
            "Full Value": "{:,.2f}",
            "Book Value": "{:,.2f}",
            "Gain/Loss": "{:,.2f}",
            "Price 100%": "{:,.4f}",
            "Clean Price": "{:,.4f}",
            "Accrued Int": "{:,.2f}",
            "Coupon": "{:.2%}",
            "YTM": "{:.2%}",
            "Yield": "{:.2%}",
            "Initial Inv Date": "{:%Y-%m-%d}",
            "Maturity Date": "{:%Y-%m-%d}",
            "Deal No.": "{:,.0f}",
        }

        # TAB 1
        with tab_holdings:
            col_search, col_dl = st.columns([3, 1])
            with col_search:
                isin_filter = st.multiselect(
                    "Filter by ISIN", options=sorted(valued["ISIN"].unique()), default=[])
            disp = valued[valued["ISIN"].isin(isin_filter)] if isin_filter else valued

            # Color Gain/Loss
            def color_gl(v):
                if isinstance(v, float):
                    return "color: #10B981" if v >= 0 else "color: #F43F5E"
                return ""

            styled = disp.style.format(fmt, na_rep="—").applymap(color_gl, subset=["Gain/Loss"])
            st.dataframe(styled, use_container_width=True, height=400)

            with col_dl:
                st.download_button(
                    "📥 Export CSV",
                    data=disp.to_csv(index=False).encode("utf-8"),
                    file_name="portfolio.csv", mime="text/csv",
                )

        # TAB 2
        with tab_scenario:
            sc1, sc2 = st.columns([1, 3])
            with sc1:
                shock_pct = st.number_input(
                    "Parallel Yield Shift (bps)", min_value=-1000, max_value=1000,
                    value=0, step=5,
                    help="e.g. 50 = +50 bps parallel shift",
                )
            shock_rate = shock_pct / 10000.0
            cleaned_shocked = cleaned.copy()
            cleaned_shocked["Yield"] = cleaned_shocked["Yield"].fillna(0) + shock_rate
            valued_shocked = run_portfolio_valuation(cleaned_shocked, valuation_timestamp)
            total_full_shocked = float(valued_shocked["Full Value"].sum())
            total_gl_shocked = float(valued_shocked["Gain/Loss"].sum())

            sm1, sm2, sm3 = st.columns(3)
            sm1.metric("Shocked Full Value", fmt_num(total_full_shocked),
                       delta=f"{total_full_shocked - total_full:,.2f}")
            sm2.metric("Shocked Gain/Loss", fmt_num(total_gl_shocked),
                       delta=f"{total_gl_shocked - total_gl:,.2f}")
            sm3.metric("Value Change %",
                       f"{((total_full_shocked - total_full) / total_full * 100):.3f}%" if total_full else "—")

            st.plotly_chart(sensitivity_curve(cleaned, valuation_timestamp),
                            use_container_width=True, config={"displayModeBar": False})

            st.dataframe(valued_shocked.style.format(fmt, na_rep="—"), use_container_width=True, height=350)

        # TAB 3
        with tab_deepdive:
            isin_options = sorted(cleaned["ISIN"].unique().tolist())
            sel_isin = st.selectbox("Select ISIN", isin_options)
            if sel_isin:
                isin_rows = valued[valued["ISIN"] == sel_isin].copy()
                total_face = float(isin_rows["Maturity Value"].sum())
                total_accrued = float(isin_rows["Accrued Int"].sum())
                total_clean_isin = float(isin_rows["Clean Value"].sum())
                total_gl_isin = float(isin_rows["Gain/Loss"].sum())

                d1, d2, d3, d4 = st.columns(4)
                d1.metric("Face Value", fmt_num(total_face))
                d2.metric("Accrued Interest", fmt_num(total_accrued))
                d3.metric("Clean Value", fmt_num(total_clean_isin))
                gl_sign2 = "+" if total_gl_isin >= 0 else ""
                d4.metric("Gain / Loss", f"{gl_sign2}{fmt_num(total_gl_isin)}")

                st.markdown("#### Deals")
                deals = isin_rows["Deal No."].astype(str).tolist()
                sel_deal = st.selectbox("Select Deal No.", deals)

                try:
                    import coupon_date
                    sel_row = cleaned[
                        (cleaned["ISIN"] == sel_isin) &
                        (cleaned["Deal No."].astype(str) == sel_deal)
                    ].iloc[0]
                    st.session_state["selected_bond"] = sel_row
                    st.markdown(f"**Lifecycle — Deal {sel_deal}**")
                    coupon_date.show_deep_dive(sel_row, valuation_timestamp)
                except Exception:
                    st.info("Detailed coupon schedules require the `coupon_date` module.")

    except Exception as e:
        st.error(f"Valuation error: {e}")

if __name__ == "__main__":
    main()